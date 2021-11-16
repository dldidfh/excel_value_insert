from config import *
from line_info import LINE_INFO_LIST
from openpyxl import load_workbook, Workbook


class ExcelControl():
    def __init__(self) -> None:
        self.init_excel_path = INIT_EXCEL_PATH
        self.sheet_name_list = EXCEL_SHEET_NAEMS

        # 형식을 가지고 있는 엑셀 로드 
        self.load_excel_sheet()


    # 저장되어있는 엑셀 형식 읽기 
    def load_excel_sheet(self):
        self.loaded_workbook = load_workbook( self.init_excel_path )
        self.loaded_sheet = self.loaded_workbook['Sheet1']

    # 엑셀 시트 생성 
    def copy_new_sheets(self, camera_num, day_key_word):
        location_name = LINE_INFO_LIST[camera_num][0]
        if day_key_word == "Friday":
            location_name_day = location_name + " 금요일"
        elif day_key_word == "Sunday":
            location_name_day = location_name + " 일요일"
        
        Sheet1 = self.loaded_workbook.get_sheet_by_name("Sheet1")
        # 만약 이미 해당 시트가 있으면 해당 시트 리턴 
        try :
            # 만약 해당 시트가 있으면 그대로 
            copy_sheet = self.loaded_workbook.get_sheet_by_name(location_name_day)
        except:
            # 만약 해당 시트가 없으면 해당 시트 리턴
            copy_sheet = self.loaded_workbook.copy_worksheet(Sheet1)

            copy_sheet.title = location_name_day
            copy_sheet['C1'] = location_name

        # for i, sheet_name in enumerate(self.sheet_name_list):
        #     if i == 0 :
        #         current_sheet = self.loaded_workbook.active
        #         current_sheet.title=sheet_name
        #         current_sheet['C1']=sheet_name
        #     else:
        #         copy_sheet = self.loaded_workbook.copy_worksheet(self.loaded_sheet)
        #         copy_sheet.title = sheet_name
        #         copy_sheet['C1'] = sheet_name
        return location_name_day


    def write_excel(self):
        self.loaded_workbook.save('output/test.xlsx')
        # new_sheet = self.loaded_workbook.create_sheet(title=sheet_name)
        # return new_sheet